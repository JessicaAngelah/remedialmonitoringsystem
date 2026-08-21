"""Tasks page — filterable table/kanban views, task creation, status updates,
and the per-subtask detail page (fields, activity/comments, attachments)."""
import csv
import io
import os

from flask import (
    Blueprint, abort, flash, redirect, render_template,
    request, send_file, send_from_directory, url_for,
)
from werkzeug.utils import secure_filename

from data import (
    PRIORITY_META, PROJECTS, STAFF, STATUS_META, TASKS, TODAY,
    assignable_people, is_overdue, log_activity, next_attachment_id, next_comment_id,
    next_task_id, notify, now_stamp, parse_date, project_of, task_of,
)
from views.core import can_edit_task, is_manager_or_admin, prefs

tasks_bp = Blueprint("tasks", __name__, url_prefix="/tasks")

UPLOAD_ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")


def _task_upload_dir(tid):
    path = os.path.join(UPLOAD_ROOT, tid)
    os.makedirs(path, exist_ok=True)
    return path


def _get_task_or_404(tid):
    tk = task_of(tid)
    if not tk:
        abort(404)
    return tk


def _filtered_sorted_tasks(q, f_project, f_assignee, f_priority, sort):
    filtered = []
    for tk in TASKS:
        if f_project != "all" and tk["projectId"] != f_project:
            continue
        if f_assignee != "all" and tk["assignee"] != f_assignee:
            continue
        if f_priority == "overdue":
            if not is_overdue(tk):
                continue
        elif f_priority != "all" and tk["priority"] != f_priority:
            continue
        if q and q.lower() not in tk["title"].lower():
            continue
        filtered.append(tk)

    if sort == "az":
        filtered.sort(key=lambda tk: tk["title"].lower())
    elif sort == "newest":
        # TASKS is appended to as tasks are created, so its list order is
        # already oldest-first — reverse that order to get newest-first
        # without needing a separate "created at" field on each task.
        task_order = {tk["id"]: i for i, tk in enumerate(TASKS)}
        filtered.sort(key=lambda tk: task_order[tk["id"]], reverse=True)
    elif sort == "deadline":
        filtered.sort(key=lambda tk: parse_date(tk["dueDate"]))

    return filtered


@tasks_bp.route("")
def tasks_view():
    q = request.args.get("q", "")
    f_project = request.args.get("project", "all")
    # Staff default to seeing only their own tasks; they can still switch
    # to "All assignees" — everyone else defaults to seeing everything.
    default_assignee = prefs()["staff_name"] if prefs()["role"] == "Staff" else "all"
    f_assignee = request.args.get("assignee", default_assignee)
    f_priority = request.args.get("priority", "all")
    sort = request.args.get("sort", "default")
    layout = request.args.get("layout", "table")
    modal = request.args.get("modal")
    edit_id = request.args.get("edit_id")
    edit_task = task_of(edit_id) if modal == "edit-task" and edit_id else None

    filtered = _filtered_sorted_tasks(q, f_project, f_assignee, f_priority, sort)

    rows = [{"task": tk, "project": project_of(tk["projectId"]), "overdue": is_overdue(tk)} for tk in filtered]

    kanban_cols = []
    if layout == "kanban":
        # "Overdue" is shown as its own column ahead of Completed/Cancelled:
        # it groups every task that reads as overdue (see is_overdue —
        # either the status field is literally "overdue", or a still-open
        # pending/in-progress task's due date has passed), regardless of
        # which of those two ways it got there. Completed/cancelled tasks
        # can never be overdue, so those columns are unaffected.
        for key in ("pending", "in-progress", "overdue", "completed", "cancelled"):
            if key == "overdue":
                col_tasks = [tk for tk in filtered if is_overdue(tk)]
            else:
                col_tasks = [tk for tk in filtered if tk["status"] == key and not is_overdue(tk)]
            kanban_cols.append({
                "key": key, "meta": STATUS_META[key],
                "tasks": [{"task": tk, "project": project_of(tk["projectId"]), "overdue": is_overdue(tk)} for tk in col_tasks],
            })

    return render_template(
        "tasks.html", view="tasks", rows=rows, total=len(TASKS), filtered_count=len(filtered),
        q=q, f_project=f_project, f_assignee=f_assignee, f_priority=f_priority, sort=sort,
        layout=layout, kanban_cols=kanban_cols, modal=modal, edit_task=edit_task,
    )


@tasks_bp.get("/export")
def export_tasks_excel():
    """Excel download of the Tasks list — honors whatever filters/sort are
    currently applied (same query params as the page itself), and always
    reads straight from the live TASKS list, so any edit made on a task's
    detail page (title, status, dates, notes, etc.) is already reflected
    the moment this is downloaded — there's no separate copy to fall out
    of sync."""
    q = request.args.get("q", "")
    f_project = request.args.get("project", "all")
    default_assignee = prefs()["staff_name"] if prefs()["role"] == "Staff" else "all"
    f_assignee = request.args.get("assignee", default_assignee)
    f_priority = request.args.get("priority", "all")
    sort = request.args.get("sort", "default")

    filtered = _filtered_sorted_tasks(q, f_project, f_assignee, f_priority, sort)

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    header_fill = PatternFill(start_color="FF004BFF", end_color="FF004BFF", fill_type="solid")
    stripe_fill = PatternFill(start_color="FFF3F8FF", end_color="FFF3F8FF", fill_type="solid")

    headers = ["Title", "Project", "Type", "Assignee", "Priority", "Status", "Overdue",
               "Start Date", "Due Date", "Progress", "Notes", "Comments"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    row_idx = 1
    for tk in filtered:
        row_idx += 1
        project = project_of(tk["projectId"])
        # Chronological (oldest first), each line stamped like a chat log —
        # e.g. "[09/08/2025, 13.05] Jafar: ..." — so the discussion reads
        # top-to-bottom the same way it happened.
        comments = sorted(tk.get("comments", []), key=lambda c: c["id"])
        comments_text = "\n".join(f"[{c.get('time', '')}] {c['author']}: {c['text']}" for c in comments)
        ws.append([
            tk["title"], project["name"] if project else "", tk.get("type", ""),
            tk.get("assignee", ""), PRIORITY_META.get(tk["priority"], {}).get("label", tk["priority"]),
            STATUS_META.get(tk["status"], {}).get("label", tk["status"]),
            "Yes" if is_overdue(tk) else "No",
            tk.get("startDate", ""), tk.get("dueDate", ""), tk.get("progress", 0),
            tk.get("notes", ""), comments_text,
        ])
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.alignment = Alignment(vertical="top", wrap_text=(col in (11, 12)))
            if row_idx % 2 == 0:
                c.fill = stripe_fill

    widths = [30, 22, 14, 16, 12, 14, 10, 13, 13, 10, 30, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name="Tasks.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@tasks_bp.post("/new")
def new_task():
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can create tasks")
        return redirect(request.form.get("redirect_to") or url_for("tasks.tasks_view"))
    title = request.form.get("title", "").strip()
    project_id = request.form.get("projectId")
    assignee = request.form.get("assignee")
    priority = request.form.get("priority", "medium")
    due_date = request.form.get("dueDate", "2026-08-10")
    ttype = request.form.get("type", "Review")
    redirect_to = request.form.get("redirect_to") or url_for("tasks.tasks_view")
    if title and project_id:
        new_id = next_task_id()
        TASKS.append({
            "id": new_id, "title": title, "projectId": project_id, "mainTaskId": None,
            "assignee": assignee, "priority": priority, "dueDate": due_date,
            "type": ttype, "status": "pending",
            "startDate": request.form.get("startDate", due_date), "progress": 0,
            "notes": "", "comments": [], "attachments": [],
        })
        flash("Task created")
        actor = prefs()["staff_name"]
        project = project_of(project_id)
        proj_txt = f" in project: {project['name']}" if project else ""
        log_activity(actor, f"{actor} created task: {title}{proj_txt}", icon="team", project_id=project_id, task_id=new_id)
    return redirect(redirect_to)


def _txt(v):
    if v is None:
        return ""
    return str(v).strip()


def _match_project(raw):
    """Match an imported project reference by exact code, then exact name
    (both case-insensitive), so 'ACME-01' or 'Acme Apparel' in the sheet
    both resolve to the right project."""
    name = _txt(raw).lower()
    if not name:
        return None
    for p in PROJECTS:
        if p["code"].lower() == name:
            return p
    for p in PROJECTS:
        if p["name"].lower() == name:
            return p
    return None


def _match_assignee(raw):
    """Match an imported name to the canonical Staff/Manager name, same
    normalization used by the project importer, so it lines up with the
    assignee dropdown instead of sitting there as an unmatched string."""
    name = _txt(raw)
    if not name:
        return ""
    for person in assignable_people():
        if person.lower() == name.lower():
            return person
    return name


def _match_date(raw):
    from datetime import date, datetime
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = _txt(raw)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


_IMPORT_TASK_COLUMNS = ["title", "project", "assignee", "priority", "status", "start date", "due date", "type", "notes"]


def _read_task_import_rows(filename, stream):
    """Load an uploaded .xlsx/.csv file into a list of {lowercased header:
    value} dicts, mirroring the project importer's approach."""
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        text = stream.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]
    if lower_name.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(stream, data_only=True)
        ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb.active
        it = ws.iter_rows()
        header = [str((c.value or "")).strip().lower() for c in next(it)]
        rows = []
        for raw in it:
            if not any(c.value for c in raw):
                continue
            row = {}
            for i, h in enumerate(header):
                if i < len(raw):
                    row[h] = raw[i].value
            rows.append(row)
        return rows
    raise ValueError("Unsupported file type — use .xlsx or .csv")


def _do_import_tasks(filename, stream):
    try:
        rows = _read_task_import_rows(filename, stream)
    except Exception as ex:
        return f"Could not read that file — {ex}"

    added, skipped = 0, 0
    for row in rows:
        title = _txt(row.get("title"))
        project = _match_project(row.get("project"))
        if not title or not project:
            skipped += 1
            continue
        priority = _txt(row.get("priority")).lower()
        if priority not in PRIORITY_META:
            priority = "medium"
        status = _txt(row.get("status")).lower().replace(" ", "-")
        if status not in STATUS_META:
            status = "pending"
        due_date = _match_date(row.get("due date")) or TODAY.isoformat()
        start_date = _match_date(row.get("start date")) or due_date
        TASKS.append({
            "id": next_task_id(), "title": title, "projectId": project["id"], "mainTaskId": None,
            "assignee": _match_assignee(row.get("assignee")), "priority": priority, "dueDate": due_date,
            "type": _txt(row.get("type")) or "Review", "status": status,
            "startDate": start_date, "progress": 100 if status == "completed" else 0,
            "notes": _txt(row.get("notes")), "comments": [], "attachments": [],
        })
        added += 1

    if added == 0:
        return "No valid rows found — check that Title and Project columns are present and Project matches an existing project's name or code"
    msg = f"Imported {added} task(s)"
    if skipped:
        msg += f" ({skipped} row(s) skipped — missing title or an unrecognized project)"
    return msg


@tasks_bp.post("/import")
def import_tasks():
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can import tasks")
        return redirect(request.form.get("redirect_to") or url_for("tasks.tasks_view"))

    redirect_to = request.form.get("redirect_to") or url_for("tasks.tasks_view")
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Choose a file to import")
        return redirect(redirect_to)

    msg = _do_import_tasks(file.filename, file)
    flash(msg)
    if msg.startswith("Imported"):
        actor = prefs()["staff_name"]
        # Rows can land on tasks across several different projects, so
        # there's no single task/project to link to — send the click to
        # the tasks list itself instead.
        log_activity(actor, f"{actor} imported tasks from {file.filename}", icon="import", url=url_for("tasks.tasks_view"))
    return redirect(redirect_to)


@tasks_bp.get("/import/template")
def download_import_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    headers = ["Title", "Project", "Assignee", "Priority", "Status", "Start Date", "Due Date", "Type", "Notes"]
    ws.append(headers)
    ws.append([
        "Reconcile appraisal figures", PROJECTS[0]["name"] if PROJECTS else "Project name or code",
        STAFF[0] if STAFF else "", "medium", "pending", "2026-08-10", "2026-08-17", "Review", "",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name="task_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@tasks_bp.post("/<tid>/delete")
def delete_task(tid):
    redirect_to = request.form.get("redirect_to") or url_for("tasks.tasks_view")
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can delete tasks")
        return redirect(redirect_to)
    tk = task_of(tid)
    TASKS[:] = [t for t in TASKS if t["id"] != tid]
    flash("Task deleted")
    if tk:
        actor = prefs()["staff_name"]
        project = project_of(tk["projectId"])
        proj_txt = f" in project: {project['name']}" if project else ""
        log_activity(actor, f"{actor} deleted task: {tk['title']}{proj_txt}", icon="delete", project_id=tk["projectId"])
    return redirect(redirect_to)


def _log_status_change(actor, tk, old_status, new_status):
    if old_status == new_status:
        return
    project = project_of(tk["projectId"])
    proj_txt = f" in project: {project['name']}" if project else ""
    if new_status == "completed":
        text = f"{actor} finished task: {tk['title']}{proj_txt}"
        icon = "done"
    else:
        label = STATUS_META.get(new_status, {}).get("label", new_status)
        text = f"{actor} changed status of task: {tk['title']} to {label}{proj_txt}"
        icon = "clock"
    log_activity(actor, text, icon=icon, project_id=tk["projectId"], task_id=tk["id"])


@tasks_bp.post("/<tid>/status")
def update_task_status(tid):
    redirect_to = request.form.get("redirect_to") or url_for("tasks.tasks_view")
    tk = task_of(tid)
    if not can_edit_task(tk):
        flash("You can only update the status of tasks assigned to you")
        return redirect(redirect_to)
    status = request.form.get("status")
    if status in STATUS_META:
        old_status = tk["status"]
        tk["status"] = status
        _log_status_change(prefs()["staff_name"], tk, old_status, status)
    return redirect(redirect_to)


# ---------------------------------------------------------------------------
# Task detail — fields, activity/comments, attachments
# ---------------------------------------------------------------------------

@tasks_bp.post("/<tid>/notify")
def notify_task(tid):
    redirect_to = request.form.get("redirect_to") or url_for("tasks.tasks_view")
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can send notifications")
        return redirect(redirect_to)
    tk = _get_task_or_404(tid)
    recipient = tk.get("assignee")
    if not recipient:
        flash("This task isn't assigned to anyone yet")
        return redirect(redirect_to)

    sender = prefs()["staff_name"]
    notify(
        recipient,
        f"{sender} sent you a reminder about task \"{tk['title']}\"",
        icon="alert",
    )
    log_activity(sender, f"{sender} notified {recipient} about task: {tk['title']}", icon="bell", project_id=tk["projectId"], task_id=tk["id"])
    flash(f"Notification sent to {recipient}")
    return redirect(redirect_to)


@tasks_bp.route("/<tid>")
def task_detail(tid):
    tk = _get_task_or_404(tid)
    project = project_of(tk["projectId"])
    comments = sorted(tk["comments"], key=lambda c: c["id"], reverse=True)
    return render_template(
        "task_detail.html", view="tasks", task=tk, project=project,
        comments=comments, attachments=tk["attachments"], STAFF=STAFF,
        is_overdue_task=is_overdue(tk),
    )


@tasks_bp.post("/<tid>/update")
def update_task_detail(tid):
    tk = _get_task_or_404(tid)
    if not can_edit_task(tk):
        flash("You can only edit tasks assigned to you")
        return redirect(url_for("tasks.task_detail", tid=tid))

    status = request.form.get("status")
    priority = request.form.get("priority")
    assignee = request.form.get("assignee", "")
    due_date = request.form.get("dueDate", "").strip()
    start_date = request.form.get("startDate", "").strip()
    progress = request.form.get("progress", "").strip()
    notes = request.form.get("notes", "")
    title = request.form.get("title", "").strip()
    ttype = request.form.get("type", "").strip()
    project_id = request.form.get("projectId", "").strip()

    old_status = tk["status"]
    if title:
        tk["title"] = title
    if ttype:
        tk["type"] = ttype
    if project_id and project_of(project_id):
        tk["projectId"] = project_id
    if status in STATUS_META:
        tk["status"] = status
    if priority in PRIORITY_META:
        tk["priority"] = priority
    tk["assignee"] = assignee  # "" means Unassigned
    if due_date:
        tk["dueDate"] = due_date
    if start_date:
        tk["startDate"] = start_date
    if progress:
        try:
            tk["progress"] = max(0, min(100, int(progress)))
        except ValueError:
            pass
    tk["notes"] = notes

    actor = prefs()["staff_name"]
    if status in STATUS_META:
        _log_status_change(actor, tk, old_status, tk["status"])

    flash("Task details saved")
    return redirect(request.form.get("redirect_to") or url_for("tasks.task_detail", tid=tid))


@tasks_bp.post("/<tid>/comments")
def add_comment(tid):
    tk = _get_task_or_404(tid)
    text = request.form.get("text", "").strip()
    if text:
        tk["comments"].append({
            "id": next_comment_id(),
            "author": prefs()["staff_name"],
            "text": text,
            "time": now_stamp(),
        })
        actor = prefs()["staff_name"]
        # Trimmed preview, not the full comment, so a long comment doesn't
        # blow out the activity feed's one-line row.
        preview = text if len(text) <= 60 else text[:57].rstrip() + "…"
        log_activity(
            actor, f'{actor} commented on task: {tk["title"]}: "{preview}"',
            icon="comment", project_id=tk["projectId"], task_id=tid,
        )
    return redirect(url_for("tasks.task_detail", tid=tid))


@tasks_bp.post("/<tid>/attachments")
def upload_attachment(tid):
    tk = _get_task_or_404(tid)
    if not can_edit_task(tk):
        flash("You can only upload documents to tasks assigned to you")
        return redirect(url_for("tasks.task_detail", tid=tid))
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Choose a file to upload")
        return redirect(url_for("tasks.task_detail", tid=tid))

    aid = next_attachment_id()
    original_name = file.filename
    safe_name = secure_filename(original_name) or "file"
    stored_name = f"{aid}__{safe_name}"
    file.save(os.path.join(_task_upload_dir(tid), stored_name))
    size = os.path.getsize(os.path.join(_task_upload_dir(tid), stored_name))

    tk["attachments"].append({
        "id": aid, "name": original_name, "storedName": stored_name,
        "size": size, "uploadedBy": prefs()["staff_name"], "time": "Just now",
    })
    flash("File uploaded")
    return redirect(url_for("tasks.task_detail", tid=tid))


@tasks_bp.get("/<tid>/attachments/<aid>/download")
def download_attachment(tid, aid):
    tk = _get_task_or_404(tid)
    att = next((a for a in tk["attachments"] if a["id"] == aid), None)
    if not att:
        abort(404)
    return send_from_directory(_task_upload_dir(tid), att["storedName"], as_attachment=True, download_name=att["name"])


@tasks_bp.post("/<tid>/attachments/<aid>/delete")
def delete_attachment(tid, aid):
    tk = _get_task_or_404(tid)
    if not can_edit_task(tk):
        flash("You can only remove documents on tasks assigned to you")
        return redirect(url_for("tasks.task_detail", tid=tid))
    att = next((a for a in tk["attachments"] if a["id"] == aid), None)
    if att:
        try:
            os.remove(os.path.join(_task_upload_dir(tid), att["storedName"]))
        except OSError:
            pass
        tk["attachments"][:] = [a for a in tk["attachments"] if a["id"] != aid]
        flash("File removed")
    return redirect(url_for("tasks.task_detail", tid=tid))
