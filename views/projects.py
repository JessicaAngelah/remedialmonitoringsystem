"""Projects pages — project grid, project detail with grouped tasks."""
import csv
import io
import os
import re
from datetime import date, datetime, timedelta

from flask import Blueprint, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, url_for

from data import (
    BUILTIN_FOLDERS, CUSTOM_FOLDERS, IMPORTS, MAIN_TASKS, PRIORITY_META, PROJECTS, STATUS_META, TASKS, TODAY,
    UPLOAD_DIR, assignable_people, is_done, is_overdue, log_activity, next_folder_id, next_project_id,
    next_task_id, next_import_id, notify, project_of, task_of,
)
from views.core import can_edit_project, is_assigned_to_project, is_manager_or_admin, prefs

projects_bp = Blueprint("projects", __name__, url_prefix="/projects")

BUILTIN_FOLDER_IDS = ("master", "weekly", "manual")


def _builtin_folder_filters():
    """(id, label) pairs for the three built-in folders, reflecting any
    renames — so every dropdown/lookup stays in sync with the ⋯ menu."""
    return [("all", "All projects")] + [(k, BUILTIN_FOLDERS[k]["name"]) for k in BUILTIN_FOLDER_IDS]

PALETTE = ["#FF003C", "#D6C100", "#004BFF", "#00FFA6"]

# Urgency -> accent color, so the case's risk level is visible at a glance
# on the project grid without reading the fine print.
URGENCY_COLORS = {
    "high": "#FF003C",
    "medium": "#D6C100",
    "low": "#00FFA6",
    "monitor": "#004BFF",
}

# Columns from the "Master" NPL sheet (debtor-per-row) mapped onto the
# project fields we store. Keys are the lower-cased header text.
_DEBTOR_FIELD_MAP = {
    "urgency": "urgency",
    "contactable": "contactable",
    "status": "npl_status",
    "remark": "remark",
    "type": "loan_type",
    "source": "source",
    "tanggal bast": "tanggal_bast",
    "pokok bast": "pokok_bast",
    "pokok": "pokok",
    "total tagihan": "total_tagihan",
    "collected 2025": "collected_2025",
    "collected 2026": "collected_2026",
    "lv agunan": "lv_agunan",
    "mv agunan": "mv_agunan",
    "jumlah aset": "jumlah_aset",
    "jaminan": "jaminan",
    "auction status": "auction_status",
    "coll": "coll",
    "dpd": "dpd",
    "due date": "due_date",
    "collector": "collector",
    "pic": "pic",
    "wo": "wo_stage",
    "wo date": "wo_date",
    "doc credit": "doc_credit",
    "profiling": "profiling",
}
_DATE_FIELDS = {"tanggal_bast", "due_date", "wo_date"}
_NUMBER_FIELDS = {"pokok_bast", "pokok", "total_tagihan", "collected_2025",
                   "collected_2026", "lv_agunan", "mv_agunan"}
# Marker appended to a header key in a parsed row to carry that cell's
# hyperlink target (see _do_import_master) alongside its display text.
_LINK_KEY_SUFFIX = "\x00link"
# Which debtor fields are document references, so an underlying hyperlink
# on the cell (e.g. "Profiling ....xlsx" linking out to Google Drive) is
# worth keeping — most fields are plain data, so only these look for one.
_LINK_FIELDS = {"profiling", "doc_credit"}


def _txt(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.upper() in ("-", "N/A", "NA") else s


def _num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = _txt(v).replace(",", "").replace("Rp", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _dt(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = _txt(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _match_person(raw):
    """Match an imported name (e.g. 'ANDRE', 'jafar') to the canonical
    Staff/Manager name ('Andre', 'Jafar') so it lines up with the
    assignee dropdown and dashboard breakdown instead of sitting there
    as a look-alike string that never matches anyone."""
    name = _txt(raw)
    if not name:
        return ""
    for person in assignable_people():
        if person.lower() == name.lower():
            return person
    return name.title()


def _make_code(name, used_codes):
    """AMRA APPAREL INDUSTRIES -> AAI-01, bumping the number on collision."""
    words = re.findall(r"[A-Za-z0-9]+", name.upper())
    base = "".join(w[0] for w in words[:4])[:4] or "NPL"
    n = 1
    code = f"{base}-{n:02d}"
    while code in used_codes:
        n += 1
        code = f"{base}-{n:02d}"
    used_codes.add(code)
    return code


def _import_debtor_rows(rows):
    """One row per debtor (the 'Master' NPL sheet) -> one project each,
    carrying over the case's financial/status fields for display."""
    used_codes = {p["code"] for p in PROJECTS}
    added = 0
    for i, row in enumerate(rows):
        name = _txt(row.get("debtor name"))
        if not name:
            continue
        urgency = _txt(row.get("urgency"))
        category = _txt(row.get("source")) or _txt(row.get("type")) or "General"
        project = {
            "id": next_project_id(),
            "code": _make_code(name, used_codes),
            "name": name,
            "category": category,
            "color": URGENCY_COLORS.get(urgency.lower(), PALETTE[i % len(PALETTE)]),
            "debtor": True,
            "origin": "master",
        }
        for header, field in _DEBTOR_FIELD_MAP.items():
            raw = row.get(header)
            if field in _DATE_FIELDS:
                project[field] = _dt(raw)
            elif field in _NUMBER_FIELDS:
                project[field] = _num(raw)
            elif field == "dpd":
                n = _num(raw)
                project[field] = int(n) if n is not None else None
            elif field == "pic":
                project[field] = _match_person(raw)
            elif field == "collector":
                project[field] = _match_person(raw)
            else:
                project[field] = _txt(raw)
            if field in _LINK_FIELDS:
                link = row.get(header + _LINK_KEY_SUFFIX)
                if link:
                    project[field + "_url"] = link
        PROJECTS.append(project)
        _assign_case_task(project)
        added += 1
    return added


def _assign_case_task(project, task_index=None):
    """Make the debtor's PIC (or Collector, when there's no PIC) the
    assignee for this case: create the overall case task if it doesn't
    exist yet, or re-point it at the (possibly updated) owner on
    re-import. This is what actually puts the case in front of them on
    the Tasks page and dashboard — project['pic']/['collector'] alone
    are just display labels until a task carries them.

    Pass a (projectId, workstream) -> task index (see _import_weekly_tasks)
    when calling this in a loop, so each call doesn't re-scan all of TASKS.
    """
    owner = project.get("pic") or project.get("collector") or ""
    if task_index is not None:
        existing = task_index.get((project["id"], "case"))
    else:
        existing = next((tk for tk in TASKS if tk["projectId"] == project["id"] and tk.get("workstream") == "case"), None)
    if existing:
        existing["assignee"] = owner
        return
    due = project.get("due_date") or _next_monday()
    task = {
        "id": next_task_id(), "projectId": project["id"], "mainTaskId": None,
        "title": "Case management", "assignee": owner,
        "priority": _priority_for_urgency(project.get("urgency")),
        "dueDate": due, "type": "Case", "status": "in-progress",
        "startDate": due, "progress": 0, "notes": project.get("remark") or "",
        "comments": [], "attachments": [], "workstream": "case",
    }
    TASKS.append(task)
    if task_index is not None:
        task_index[(project["id"], "case")] = task


def _import_simple_rows(rows):
    """Generic fallback: header row of code, name, category, color."""
    added = 0
    for i, row in enumerate(rows):
        code = _txt(row.get("code"))
        name = _txt(row.get("name"))
        if not code or not name:
            continue
        category = _txt(row.get("category")) or "General"
        color = _txt(row.get("color")) or PALETTE[i % len(PALETTE)]
        PROJECTS.append({"id": next_project_id(), "code": code, "name": name, "category": category, "color": color, "origin": "master"})
        added += 1
    return added


# ---------------------------------------------------------------------------
# Tasks imported from debtor-level activity sheets ("Interaksi" in the
# Master workbook, and the weekly workstream reports). Both key off the
# debtor's name to find the matching project — names are normalized (case,
# punctuation, trailing "(code)" suffixes) so minor formatting differences
# between sheets still match.
# ---------------------------------------------------------------------------

def _norm_name(name):
    s = re.sub(r"\(.*?\)", "", name or "")
    s = re.sub(r"[^A-Za-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip().upper()


def _debtor_lookup():
    return {_norm_name(p["name"]): p for p in PROJECTS if p.get("debtor")}


def _next_monday():
    days_ahead = (7 - TODAY.weekday()) % 7 or 7
    return (TODAY + timedelta(days=days_ahead)).isoformat()


def _priority_for_urgency(urgency):
    return {"high": "high", "medium": "medium", "low": "low", "monitor": "low"}.get((urgency or "").lower(), "medium")


def _import_interaksi_tasks(ws):
    """'Interaksi' sheet = one field visit/call per row -> one task each.

    Assignee is the row's Surveyor when given, otherwise the debtor's PIC
    (or Collector, if there's no PIC) from the Master sheet — the person
    in charge / doing the work, so it's the right fallback owner for a
    visit that doesn't name one.
    """
    lookup = _debtor_lookup()
    it = ws.iter_rows(values_only=True)
    header = [str(h or "").strip().lower() for h in next(it)]

    def col(name):
        return header.index(name) if name in header else None

    c_debitur, c_surveyor, c_objektif = col("debitur"), col("surveyor"), col("objektif")
    c_tgl, c_status, c_hasil = col("tgl kunjungan"), col("status"), col("hasil")
    if c_debitur is None:
        return 0, 0

    def get(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    added, unmatched = 0, set()
    for row in it:
        if not any(row):
            continue
        name = _txt(get(row, c_debitur))
        if not name:
            continue
        project = lookup.get(_norm_name(name))
        if not project:
            unmatched.add(name)
            continue
        objektif = _txt(get(row, c_objektif)) or "Visit"
        met = _txt(get(row, c_status)).lower() == "bertemu"
        due = _dt(get(row, c_tgl)) or _next_monday()
        TASKS.append({
            "id": next_task_id(), "projectId": project["id"], "mainTaskId": None,
            "title": f"Visit — {objektif}",
            "assignee": _txt(get(row, c_surveyor)) or project.get("pic") or project.get("collector") or "",
            "priority": "medium", "dueDate": due, "type": "Field Visit",
            "status": "completed" if met else "pending",
            "startDate": due, "progress": 100 if met else 30,
            "notes": _txt(get(row, c_hasil)), "comments": [], "attachments": [],
            "workstream": "interaksi",
        })
        added += 1
    return added, len(unmatched)


# Weekly workstream report -> (workstream key, task title). Several sheets
# feed the same workstream (e.g. the standalone PKPU case sheet) so re-runs
# update the one task per debtor per workstream instead of piling up.
_WEEKLY_SHEETS = {
    "Monitoring Internal Coll": ("internal_coll", "Internal Collection follow-up"),
    "Monitoring Write Off": ("write_off", "Write-off recovery"),
    "Monitoring Auction": ("auction", "Auction / Collateral process"),
    "PKPU": ("pkpu", "PKPU proceedings"),
    "Wendra Wilfrida": ("pkpu", "PKPU proceedings"),
}
_WEEKLY_META_HEADERS = {"no", "debtor", "debtor name", "agency name", "information",
                         "tgl write-off", "noa", "overdue amount", "kol", "plan next week"}
_WORKSTREAM_CATEGORY = {
    "internal_coll": "Internal Collection",
    "write_off": "Write Off",
    "auction": "Auction",
    "pkpu": "PKPU",
}


def _parse_weekly_sheet(ws):
    """Auto-detect the header row + debtor/overdue columns, then treat every
    other labeled column (W1, W2, "JUN'26", "Plan Next Week", ...) as a
    narrative update column — the last non-empty one is the latest update.

    Reads the sheet once via iter_rows(values_only=True) instead of pulling
    cells one at a time with ws.cell(row, column) — the latter is orders of
    magnitude slower in openpyxl on anything but tiny sheets, since each
    call does a lookup + wraps the value in a Cell object. Buffering all
    rows up front lets us both hunt for the header and walk the data
    without ever touching the slow per-cell API.
    """
    all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    return _parse_weekly_rows(all_rows)


def _parse_pasted_weekly_rows(text):
    """Same row shape as _parse_weekly_sheet, but from text pasted straight
    out of Excel (tab-separated columns, one line per row) instead of a
    workbook sheet — for updating a single case without a whole workbook.
    Blank lines are dropped; each remaining line is split on tabs.
    """
    all_rows = [line.split("\t") for line in text.replace("\r\n", "\n").split("\n") if line.strip()]
    return _parse_weekly_rows(all_rows)


def _parse_weekly_rows(all_rows):
    header, header_row_idx = [], None
    for i, raw in enumerate(all_rows[:10]):
        norm = [str(v).strip().lower() if v is not None else "" for v in raw]
        if "debtor" in norm or "agency name" in norm or "debtor name" in norm:
            header, header_row_idx = norm, i
            break
    if header_row_idx is None:
        return []

    def find_col(names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    no_col = find_col({"no"})
    debtor_col = find_col({"debtor", "debtor name", "agency name"})
    overdue_col = find_col({"overdue amount"})
    kol_col = find_col({"kol"})
    plan_col = find_col({"plan next week"})
    if no_col is None or debtor_col is None:
        return []
    narrative_cols = [i for i, h in enumerate(header) if h and h not in _WEEKLY_META_HEADERS]

    results = []
    for vals in all_rows[header_row_idx + 1:]:
        if no_col >= len(vals) or _num(vals[no_col]) is None:
            continue  # skips section headers / grand-total rows, which have no serial number
        name = _txt(vals[debtor_col]) if debtor_col < len(vals) else ""
        if not name:
            continue
        notes = ""
        for i in narrative_cols:
            if i < len(vals) and vals[i] not in (None, ""):
                notes = str(vals[i]).strip()
        overdue = _num(vals[overdue_col]) if overdue_col is not None and overdue_col < len(vals) else None
        kol = _txt(vals[kol_col]) if kol_col is not None and kol_col < len(vals) else ""
        plan = _txt(vals[plan_col]) if plan_col is not None and plan_col < len(vals) else ""
        results.append((name, notes, overdue, kol, plan))
    return results


def _get_or_create_weekly_project(lookup, name, workstream, overdue, kol=""):
    """Find the debtor's existing project (usually created by a Master-sheet
    import) or, if this debtor has never been seen before, create a minimal
    project for them on the spot — so a Weekly report can be imported on
    its own instead of silently no-op'ing on every row. 'Overdue Amount' in
    the weekly sheets is in millions of Rupiah, same scale used elsewhere
    (Master's Pokok/Total Tagihan are raw Rupiah), so it's scaled up x1e6
    before being stored in total_tagihan.
    """
    key = _norm_name(name)
    project = lookup.get(key)
    if project:
        if overdue is not None:
            project["total_tagihan"] = overdue * 1_000_000
        if kol:
            project["coll"] = kol
        return project, False
    used_codes = {p["code"] for p in PROJECTS}
    project = {
        "id": next_project_id(), "code": _make_code(name, used_codes), "name": name,
        "category": _WORKSTREAM_CATEGORY.get(workstream, "General"),
        "color": PALETTE[len(PROJECTS) % len(PALETTE)],
        "debtor": True,
        "origin": "weekly",
        # A debtor may first show up in a Weekly report instead of a Master
        # import, so this project won't get filled in through _DEBTOR_FIELD_MAP.
        # Seed every NPL field the templates read (project.pokok, etc.) with
        # None so they render as "—" instead of raising on a missing key.
        **{field: None for field in _DEBTOR_FIELD_MAP.values()},
        "plan_next_week": None,
    }
    if overdue is not None:
        project["total_tagihan"] = overdue * 1_000_000
    if kol:
        project["coll"] = kol
    PROJECTS.append(project)
    lookup[key] = project
    return project, True


def _apply_weekly_rows(lookup, task_index, workstream, title, rows):
    """Create/update the one task per debtor per workstream for a batch of
    already-parsed (name, notes, overdue, kol, plan) rows — shared by the
    full workbook import and the single-case paste import below, so both
    go through the exact same match/create/update logic. Also returns
    every project id touched (in row order, duplicates included) so a
    caller that only ever sees one distinct project — the single-case
    import — can link its activity-log entry straight to it."""
    created, updated, new_projects = 0, 0, 0
    touched_ids = []
    for name, notes, overdue, kol, plan in rows:
        project, is_new = _get_or_create_weekly_project(lookup, name, workstream, overdue, kol)
        touched_ids.append(project["id"])
        if plan:
            project["plan_next_week"] = plan
        if is_new:
            new_projects += 1
            _assign_case_task(project, task_index)
        existing = task_index.get((project["id"], workstream))
        if existing:
            if notes:
                existing["notes"] = notes
            existing["dueDate"] = _next_monday()
            updated += 1
        else:
            due = _next_monday()
            task = {
                "id": next_task_id(), "projectId": project["id"], "mainTaskId": None,
                "title": title, "assignee": project.get("pic") or project.get("collector") or "",
                "priority": _priority_for_urgency(project.get("urgency")),
                "dueDate": due, "type": "Weekly Update", "status": "in-progress",
                "startDate": due, "progress": 0, "notes": notes,
                "comments": [], "attachments": [], "workstream": workstream,
            }
            TASKS.append(task)
            task_index[(project["id"], workstream)] = task
            created += 1
    return created, updated, new_projects, touched_ids


def _import_weekly_tasks(wb):
    lookup = _debtor_lookup()
    # Index existing tasks by (projectId, workstream) once up front instead
    # of scanning the whole TASKS list for every row of every sheet. TASKS
    # only grows week over week, so that repeated full-list scan is what
    # made re-imports get slower and slower over time.
    task_index = {(tk["projectId"], tk.get("workstream")): tk for tk in TASKS}
    created, updated, new_projects = 0, 0, 0
    for sheet_name, (workstream, title) in _WEEKLY_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        rows = _parse_weekly_sheet(wb[sheet_name])
        c, u, n, _touched = _apply_weekly_rows(lookup, task_index, workstream, title, rows)
        created += c
        updated += u
        new_projects += n
    return created, updated, new_projects


# Workstream choices offered on the single-case (paste) import — value must
# match the workstream keys used above so it lands in the same bucket a
# full workbook import for that sheet would.
_SINGLE_CASE_WORKSTREAMS = [
    ("pkpu", "PKPU proceedings"),
    ("internal_coll", "Internal Collection follow-up"),
    ("write_off", "Write-off recovery"),
    ("auction", "Auction / Collateral process"),
]


def _do_import_single_case(workstream, text, file_storage=None, touched_out=None):
    """Update (or create) just one debtor's case, either from a single row
    pasted straight out of Excel or from a small uploaded .xlsx file for
    that one case — same parsing/matching rules as a full Weekly workbook
    import, so there's no need to build (or paste from) a whole workbook
    just to log one case's update. The uploaded file wins if both are
    provided.

    `touched_out`, if given, is a list the caller passes in empty and this
    function appends every touched project id to — an out-param rather
    than a return-value change so every existing early "return <message>"
    line below doesn't need touching. The caller uses it to link the
    resulting activity-log entry straight to the project when there's
    exactly one (the normal case — a single pasted/uploaded row)."""
    title = dict(_SINGLE_CASE_WORKSTREAMS).get(workstream)
    if not title:
        return "Choose a workstream for this case"

    if file_storage is not None and getattr(file_storage, "filename", ""):
        if not file_storage.filename.lower().endswith(".xlsx"):
            return "The case file needs to be an .xlsx file"
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_storage.read()), data_only=True)
        except Exception as ex:
            return f"Could not read that file — {ex}"
        # Prefer a sheet named for this workstream (e.g. a one-row export
        # of "Monitoring Auction") if present, otherwise fall back to
        # whichever sheet is active — a single-case file is usually just
        # one small sheet with no particular name.
        matching_sheets = [name for name, (ws, _) in _WEEKLY_SHEETS.items() if ws == workstream]
        ws = next((wb[name] for name in matching_sheets if name in wb.sheetnames), wb.active)
        rows = _parse_weekly_sheet(ws)
        if not rows:
            return "Could not find a debtor row in that file — make sure it has a header row (No, Debtor, ...) and the case's row below it"
    else:
        text = (text or "").strip()
        if not text:
            return "Paste the case's row or upload a file for it first"
        rows = _parse_pasted_weekly_rows(text)
        if not rows:
            return "Could not find a debtor row — make sure you included the header row (No, Debtor, ...) above it"

    lookup = _debtor_lookup()
    task_index = {(tk["projectId"], tk.get("workstream")): tk for tk in TASKS}
    created, updated, new_projects, touched_ids = _apply_weekly_rows(lookup, task_index, workstream, title, rows)
    if touched_out is not None:
        touched_out.extend(touched_ids)
    names = ", ".join(r[0] for r in rows)
    msg = f"Updated case: {names}" if updated and not created else f"Added case: {names}"
    if new_projects:
        msg += " (new project created)"
    return msg

# (filter key, label, predicate over a card dict)
STATUS_FILTERS = [
    ("all", "All statuses"),
    ("not-started", "Not started"),
    ("in-progress", "In progress"),
    ("completed", "Completed"),
    ("overdue", "Overdue"),
]


def _weekly_touched_ids():
    """Project ids with at least one task created/refreshed by a Weekly
    import (any of the Internal Coll / Write Off / Auction / PKPU
    workstreams) — used so the "Weekly" folder reflects every debtor that
    showed up in the latest weekly workbook, not just the handful that
    happened to be brand new (origin=="weekly" only covers those)."""
    return {tk["projectId"] for tk in TASKS if tk.get("type") == "Weekly Update"}


def _projects_in_folder(key, plist):
    """Projects belonging to a given folder key. "weekly" is membership by
    activity (touched by a weekly import) rather than by "origin", since a
    debtor already living in Master data still needs to show up here once
    a weekly update lands for them — origin only tells us where a project
    was *first* created, not everywhere it currently belongs."""
    if key == "weekly":
        touched = _weekly_touched_ids()
        return [p for p in plist if p["id"] in touched]
    return [p for p in plist if p.get("origin", "manual") == key]


def _card_status(card):
    """Single overall status for a project card, used for filtering/badges."""
    if card["overdue"] > 0:
        return "overdue"
    if card["total"] > 0 and card["done"] == card["total"]:
        return "completed"
    if card["done"] == 0:
        return "not-started"
    return "in-progress"


@projects_bp.route("")
def projects_list():
    modal = request.args.get("modal")
    q = request.args.get("q", "").strip().lower()
    # Default to "All projects" for everyone; they can still switch to
    # "My projects" (only cases the user is PIC/collector or has a task
    # assigned on).
    f_scope = request.args.get("scope", "all")

    # Built-in folders plus any user-created ones, so custom folders show up
    # as cards right alongside Master/Weekly/Manual — but only once there's
    # actually something in them; an unused built-in category or a
    # freshly-created empty folder is just noise on this page.
    all_folders = _builtin_folder_filters() + [(f["id"], f["name"]) for f in CUSTOM_FOLDERS]

    plist = PROJECTS
    if f_scope == "mine":
        my_name = prefs()["staff_name"]
        plist = [p for p in plist if is_assigned_to_project(p, my_name)]
    if q:
        plist = [
            p for p in plist
            if q in p["name"].lower() or q in p["code"].lower() or q in p["category"].lower()
        ]

    # Most recent upload of each kind, so a folder card can offer a
    # "View excel" button straight to the original spreadsheet's contents
    # (shown inline, not downloaded) without digging through the activity log.
    latest_master = next((imp for imp in reversed(IMPORTS) if imp["kind"] == "master"), None)
    latest_weekly = next((imp for imp in reversed(IMPORTS) if imp["kind"] == "weekly"), None)
    folder_import = {"master": latest_master, "weekly": latest_weekly}

    # Aggregate task stats per folder (for the folder-card grid). Built-in
    # categories (Master/Weekly/Manual) are hidden when empty since they're
    # not real records — but a user-created folder always shows, even with
    # zero projects in it yet, so it can still be renamed or deleted.
    folder_cards = []
    custom_by_id = {f["id"]: f for f in CUSTOM_FOLDERS}
    for key, label in all_folders:
        if key == "all":
            continue
        is_custom = key not in BUILTIN_FOLDER_IDS
        f_projects = _projects_in_folder(key, plist)
        if not f_projects and not is_custom:
            continue
        f_tasks = [tk for tk in TASKS if tk["projectId"] in {p["id"] for p in f_projects}]
        f_done = len([tk for tk in f_tasks if is_done(tk)])
        f_pct = round((f_done / len(f_tasks)) * 100) if f_tasks else 0
        f_import = folder_import.get(key)
        # Every folder — built-in or custom — shows its own saved color, so
        # the ⋯ menu's swatches always line up with what's actually on the
        # card, and a rename/recolor sticks instead of shifting around
        # depending on which other folders happen to be visible.
        card_color = custom_by_id[key].get("color", "#004BFF") if is_custom else BUILTIN_FOLDERS[key]["color"]
        folder_cards.append({
            "key": key, "label": label, "projects": len(f_projects), "total": len(f_tasks),
            "done": f_done, "pct": f_pct, "color": card_color,
            "custom": is_custom, "builtin": not is_custom,
            "import_id": f_import["id"] if f_import else None,
        })

    edit_folder = None
    if modal == "edit-folder":
        fid = request.args.get("folder", "")
        if fid in BUILTIN_FOLDER_IDS:
            b = BUILTIN_FOLDERS[fid]
            edit_folder = {"key": fid, "label": b["name"], "color": b["color"], "builtin": True}
        else:
            src = next((f for f in CUSTOM_FOLDERS if f["id"] == fid), None)
            if src:
                edit_folder = {"key": src["id"], "label": src["name"], "color": src.get("color", "#004BFF"), "builtin": False}
            else:
                modal = None

    return render_template(
        "projects_list.html", view="projects", modal=modal, q=q, f_scope=f_scope,
        folder_cards=folder_cards, SINGLE_CASE_WORKSTREAMS=_SINGLE_CASE_WORKSTREAMS,
        edit_folder=edit_folder,
    )


@projects_bp.route("/folders/<folder_key>")
def folder_detail(folder_key):
    """A single folder's own page: just its projects, not every other
    folder alongside them."""
    modal = request.args.get("modal")
    q = request.args.get("q", "").strip().lower()
    f_status = request.args.get("status", "all")
    f_scope = request.args.get("scope", "all")

    all_folders = _builtin_folder_filters() + [(f["id"], f["name"]) for f in CUSTOM_FOLDERS]
    folder_labels = dict(all_folders)
    if folder_key not in folder_labels or folder_key == "all":
        flash("Folder not found")
        return redirect(url_for("projects.projects_list"))

    plist = _projects_in_folder(folder_key, PROJECTS)
    if f_scope == "mine":
        my_name = prefs()["staff_name"]
        plist = [p for p in plist if is_assigned_to_project(p, my_name)]
    if q:
        plist = [
            p for p in plist
            if q in p["name"].lower() or q in p["code"].lower() or q in p["category"].lower()
        ]

    cards = []
    for p in plist:
        pts = [tk for tk in TASKS if tk["projectId"] == p["id"]]
        pdone = len([tk for tk in pts if is_done(tk)])
        povr = len([tk for tk in pts if is_overdue(tk)])
        pct = round((pdone / len(pts)) * 100) if pts else 0
        card = {"project": p, "total": len(pts), "done": pdone, "overdue": povr, "pct": pct}
        card["status"] = _card_status(card)
        cards.append(card)

    if f_status != "all":
        cards = [c for c in cards if c["status"] == f_status]

    import_id = None
    if folder_key == "master":
        latest = next((imp for imp in reversed(IMPORTS) if imp["kind"] == "master"), None)
        import_id = latest["id"] if latest else None
    elif folder_key == "weekly":
        latest = next((imp for imp in reversed(IMPORTS) if imp["kind"] == "weekly"), None)
        import_id = latest["id"] if latest else None

    return render_template(
        "folder_detail.html", view="projects", cards=cards, modal=modal, q=q,
        f_status=f_status, STATUS_FILTERS=STATUS_FILTERS, f_scope=f_scope,
        folder_key=folder_key, folder_label=folder_labels[folder_key], import_id=import_id,
    )


@projects_bp.route("/<pid>")
def project_detail(pid):
    project = project_of(pid)
    if not project:
        flash("Project not found")
        return redirect(url_for("projects.projects_list"))
    tasks = sorted(
        (tk for tk in TASKS if tk["projectId"] == pid),
        key=lambda tk: (tk["status"] == "completed", tk["dueDate"]),
    )
    pdone = len([tk for tk in tasks if is_done(tk)])
    pct = round((pdone / len(tasks)) * 100) if tasks else 0

    modal = request.args.get("modal")
    edit_id = request.args.get("edit_id")
    edit_task = task_of(edit_id) if modal == "edit-task" and edit_id else None
    return render_template(
        "project_detail.html", view="projects", project=project, tasks=tasks,
        pdone=pdone, ptotal=len(tasks), pct=pct, modal=modal, edit_task=edit_task,
    )


def _apply_case_details(project, form):
    """Read the Case Details fields (PIC, financials, Kol, monthly notes,
    ...) from a submitted form onto a debtor-type project. Shared by manual
    project creation and the edit-project form so both stay in sync."""
    project["pic"] = _match_person(form.get("pic", ""))
    project["collector"] = _match_person(form.get("collector", ""))
    project["npl_status"] = _txt(form.get("npl_status", ""))
    project["urgency"] = _txt(form.get("urgency", ""))
    project["contactable"] = _txt(form.get("contactable", ""))
    project["loan_type"] = _txt(form.get("loan_type", ""))
    project["source"] = _txt(form.get("source", ""))
    project["jaminan"] = _txt(form.get("jaminan", ""))
    project["pokok"] = _num(form.get("pokok", ""))
    project["total_tagihan"] = _num(form.get("total_tagihan", ""))
    project["collected_2025"] = _num(form.get("collected_2025", ""))
    project["collected_2026"] = _num(form.get("collected_2026", ""))
    project["lv_agunan"] = _num(form.get("lv_agunan", ""))
    project["mv_agunan"] = _num(form.get("mv_agunan", ""))
    dpd = _num(form.get("dpd", ""))
    project["dpd"] = int(dpd) if dpd is not None else None
    project["due_date"] = _dt(form.get("due_date", ""))
    project["wo_stage"] = _txt(form.get("wo_stage", ""))
    project["wo_date"] = _dt(form.get("wo_date", ""))
    project["auction_status"] = _txt(form.get("auction_status", ""))
    project["tanggal_bast"] = _dt(form.get("tanggal_bast", ""))
    project["coll"] = _txt(form.get("kol", ""))
    project["tipe"] = _txt(form.get("tipe", ""))
    project["surveyor"] = _match_person(form.get("surveyor", ""))
    project["objektif"] = _txt(form.get("objektif", ""))
    project["meeting_status"] = _txt(form.get("meeting_status", ""))
    project["critical"] = _txt(form.get("critical", ""))
    project["rating"] = _txt(form.get("rating", ""))
    project["remark"] = _txt(form.get("remark", ""))
    project["plan_next_week"] = _txt(form.get("plan_next_week", ""))

    titles = form.getlist("custom_title")
    descs = form.getlist("custom_desc")
    project["custom_details"] = [
        {"title": _txt(t), "description": _txt(d)}
        for t, d in zip(titles, descs) if _txt(t) or _txt(d)
    ]


# ---------------------------------------------------------------------------
# Per-project "view/edit/download as Excel" — same idea as the imported
# workbook viewer (view_import/save_import_edits/download_import), but the
# sheet is generated live from a single project's own fields instead of a
# stored file, so there's nothing to re-import: edits write straight back
# onto the project, and the download is built on the fly.
#
# Fields are grouped into named sections so the editor can render them as
# tabs instead of one long scrolling table.
# ---------------------------------------------------------------------------
PROJECT_EXCEL_BASE_FIELDS = [
    ("name", "Project / debtor name"),
    ("code", "Code"),
    ("category", "Category"),
]
PROJECT_EXCEL_CASE_GROUPS = [
    ("General", [
        ("pic", "PIC"),
        ("collector", "Collector"),
        ("npl_status", "Status"),
        ("urgency", "Urgency"),
        ("contactable", "Contactable"),
        ("loan_type", "Loan type"),
        ("source", "Source"),
        ("jaminan", "Jaminan"),
        ("tipe", "Tipe"),
        ("surveyor", "Surveyor"),
        ("objektif", "Objektif"),
        ("meeting_status", "Meeting status"),
        ("critical", "Critical"),
        ("rating", "Rating"),
    ]),
    ("Financials", [
        ("pokok", "Pokok"),
        ("total_tagihan", "Overdue amount / Total tagihan"),
        ("collected_2025", "Collected 2025"),
        ("collected_2026", "Collected 2026"),
        ("lv_agunan", "LV Agunan"),
        ("mv_agunan", "MV Agunan"),
    ]),
    ("Timeline & Docs", [
        ("dpd", "DPD"),
        ("due_date", "Due date"),
        ("wo_stage", "WO stage"),
        ("wo_date", "WO date"),
        ("auction_status", "Auction status"),
        ("tanggal_bast", "Tanggal BAST"),
        ("coll", "Kol"),
        ("profiling", "Profiling"),
        ("doc_credit", "Doc Credit"),
    ]),
    ("Notes", [
        ("remark", "Remark"),
        ("plan_next_week", "Plan next week"),
    ]),
]
# Flat view of the same fields, kept for code that just needs "every case
# field" (valid-key checks, the flat Excel download, etc.) without caring
# about which tab it lives on.
PROJECT_EXCEL_CASE_FIELDS = [f for _, fields in PROJECT_EXCEL_CASE_GROUPS for f in fields]
_EXCEL_DATE_FIELDS = {"due_date", "wo_date", "tanggal_bast"}
_EXCEL_NUMBER_FIELDS = {"pokok", "total_tagihan", "collected_2025", "collected_2026", "lv_agunan", "mv_agunan"}


def _project_excel_fields(project):
    fields = list(PROJECT_EXCEL_BASE_FIELDS)
    if project.get("debtor"):
        fields += PROJECT_EXCEL_CASE_FIELDS
    return fields


def _cell_display(v, key=None):
    """Value shown in the in-app spreadsheet grid. Whole-number floats are
    shown as plain ints; currency/amount fields (_EXCEL_NUMBER_FIELDS) get
    thousands separators (1000000 -> "1,000,000") for readability — safe to
    round-trip since _num() strips commas back out on save."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if key in _EXCEL_NUMBER_FIELDS and isinstance(v, (int, float)):
        return f"{v:,.0f}" if float(v).is_integer() else f"{v:,.2f}"
    return v


def _set_project_field(project, key, raw):
    if key in ("pic", "collector", "surveyor"):
        project[key] = _match_person(raw)
    elif key in _EXCEL_DATE_FIELDS:
        project[key] = _dt(raw)
    elif key in _EXCEL_NUMBER_FIELDS:
        project[key] = _num(raw)
    elif key == "dpd":
        n = _num(raw)
        project[key] = int(n) if n is not None else None
    elif key == "code":
        val = _txt(raw).upper()
        if val:
            project[key] = val
    elif key == "name":
        val = _txt(raw)
        if val:
            project[key] = val
    elif key == "category":
        project[key] = _txt(raw) or "General"
    else:
        project[key] = _txt(raw)


def _rows_for(project, fields):
    return [
        {"key": k, "label": label, "value": _cell_display(project.get(k), key=k),
         "url": project.get(k + "_url") if k in _LINK_FIELDS else None}
        for k, label in fields
    ]


@projects_bp.get("/<pid>/excel")
def view_project_excel(pid):
    """Show the project's own fields as an editable spreadsheet, split into
    tabbed sections (Overview, General, Financials, ...) instead of one
    long scrolling table, the same way an imported workbook can be opened
    and edited in-app."""
    project = project_of(pid)
    if not project:
        flash("Project not found")
        return redirect(url_for("projects.projects_list"))
    groups = [("Overview", _rows_for(project, PROJECT_EXCEL_BASE_FIELDS))]
    if project.get("debtor"):
        groups += [(name, _rows_for(project, fields)) for name, fields in PROJECT_EXCEL_CASE_GROUPS]
    return render_template(
        "project_excel.html", view="projects", project=project, groups=groups,
        can_edit=can_edit_project(project), fixed_screen=True,
    )


@projects_bp.post("/<pid>/excel/save")
def save_project_excel(pid):
    project = project_of(pid)
    if not project:
        return jsonify({"ok": False, "error": "Project not found"}), 404
    if not can_edit_project(project):
        return jsonify({"ok": False, "error": "You can't edit this project"}), 403

    payload = request.get_json(silent=True) or {}
    edits = payload.get("edits") or []
    if not isinstance(edits, list) or not edits:
        return jsonify({"ok": False, "error": "Nothing to save"}), 400

    valid_keys = {k for k, _ in _project_excel_fields(project)}
    for e in edits:
        key, value = e.get("key"), e.get("value", "")
        if key not in valid_keys:
            continue
        _set_project_field(project, key, value)

    if project.get("debtor"):
        _assign_case_task(project)

    actor = prefs()["staff_name"]
    log_activity(actor, f"{actor} edited project via excel view: {project['name']}", icon="import", project_id=pid, url=url_for("projects.view_project_excel", pid=pid))
    return jsonify({"ok": True})


@projects_bp.get("/<pid>/excel/download")
def download_project_excel(pid):
    project = project_of(pid)
    if not project:
        flash("Project not found")
        return redirect(url_for("projects.projects_list"))

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project"

    header_fill = PatternFill(start_color="FF004BFF", end_color="FF004BFF", fill_type="solid")
    stripe_fill = PatternFill(start_color="FFF3F8FF", end_color="FFF3F8FF", fill_type="solid")

    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    row_idx = 1
    for k, label in _project_excel_fields(project):
        row_idx += 1
        raw = project.get(k)
        value = int(raw) if isinstance(raw, float) and raw.is_integer() else raw
        ws.append([label, value if value is not None else ""])
        label_cell, value_cell = ws.cell(row=row_idx, column=1), ws.cell(row=row_idx, column=2)
        # Real Excel number formatting (not a pre-formatted string) so the
        # thousands separators show up while the cell stays a usable number
        # for anyone who wants to sum/reference it — e.g. 1000000 -> 1,000,000.
        if k in _EXCEL_NUMBER_FIELDS and isinstance(value, (int, float)):
            value_cell.number_format = "#,##0"
        # Alternating row shading, skipping the header row, for readability.
        if row_idx % 2 == 0:
            label_cell.fill = stripe_fill
            value_cell.fill = stripe_fill

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A2"

    # A second sheet lists every task under the project — including the
    # comment thread — so notes/discussion logged in the app aren't lost
    # when the case is exported. Comments are chronological (oldest
    # first) and stamped like a chat log, e.g.
    # "[09/08/2025, 13.05] Jafar: ...", so the thread reads top-to-bottom
    # the same way it happened.
    project_tasks = [tk for tk in TASKS if tk["projectId"] == pid]
    if project_tasks:
        ws2 = wb.create_sheet("Tasks")
        task_headers = ["Title", "Type", "Assignee", "Priority", "Status", "Start Date", "Due Date", "Progress", "Notes", "Comments"]
        ws2.append(task_headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        t_row = 1
        for tk in project_tasks:
            t_row += 1
            comments = sorted(tk.get("comments", []), key=lambda c: c["id"])
            comments_text = "\n".join(f"[{c.get('time', '')}] {c['author']}: {c['text']}" for c in comments)
            ws2.append([
                tk["title"], tk.get("type", ""), tk.get("assignee", ""),
                PRIORITY_META.get(tk["priority"], {}).get("label", tk["priority"]),
                STATUS_META.get(tk["status"], {}).get("label", tk["status"]),
                tk.get("startDate", ""), tk.get("dueDate", ""), tk.get("progress", 0),
                tk.get("notes", ""), comments_text,
            ])
            for col in range(1, len(task_headers) + 1):
                c = ws2.cell(row=t_row, column=col)
                c.alignment = Alignment(vertical="top", wrap_text=(col in (9, 10)))
                if t_row % 2 == 0:
                    c.fill = stripe_fill

        widths = [28, 14, 16, 12, 14, 13, 13, 10, 30, 42]
        for i, w in enumerate(widths, start=1):
            ws2.column_dimensions[chr(64 + i)].width = w
        ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = re.sub(r"[\\/:*?\"<>|]", "-", f"{project['code']} - {project['name']}")
    return send_file(
        buf, as_attachment=True, download_name=f"{safe_name}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@projects_bp.post("/new")
def new_project():
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can create projects")
        return redirect(url_for("projects.projects_list"))
    name = request.form.get("name", "").strip()
    code = request.form.get("code", "").strip().upper()
    category = request.form.get("category", "").strip() or "General"
    color = request.form.get("color", "#004BFF")

    folder_choice = request.form.get("folder_choice", "manual")
    if folder_choice == "new":
        new_name = request.form.get("new_folder_name", "").strip()
        if new_name:
            fid = next_folder_id()
            CUSTOM_FOLDERS.append({"id": fid, "name": new_name, "color": color})
            folder = fid
            actor = prefs()["staff_name"]
            log_activity(actor, f"{actor} created folder: {new_name}", icon="folder", url=url_for("projects.folder_detail", folder_key=fid))
        else:
            folder = "manual"
    elif folder_choice == "existing":
        folder = request.form.get("folder", "manual")
        valid_folder_ids = {"master", "weekly", "manual"} | {f["id"] for f in CUSTOM_FOLDERS}
        if folder not in valid_folder_ids:
            folder = "manual"
    else:
        folder = "manual"

    if name:
        if not code:
            used_codes = {p["code"] for p in PROJECTS}
            code = _make_code(name, used_codes)
        pid = next_project_id()
        project = {"id": pid, "name": name, "code": code, "category": category, "color": color, "origin": folder}
        if request.form.get("is_debtor") == "on":
            project["debtor"] = True
            for field in _DEBTOR_FIELD_MAP.values():
                project.setdefault(field, None)
            _apply_case_details(project, request.form)
            PROJECTS.append(project)
            _assign_case_task(project)
        else:
            PROJECTS.append(project)
        flash("Project created")
        actor = prefs()["staff_name"]
        log_activity(actor, f"{actor} created project: {name}", icon="team", project_id=pid)
    return redirect(url_for("projects.projects_list"))


@projects_bp.post("/folders/new")
def new_folder():
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can create folders")
        return redirect(url_for("projects.projects_list"))
    name = request.form.get("name", "").strip()
    color = request.form.get("color", "#004BFF")
    if name:
        fid = next_folder_id()
        CUSTOM_FOLDERS.append({"id": fid, "name": name, "color": color})
        flash("Folder created")
        actor = prefs()["staff_name"]
        log_activity(actor, f"{actor} created folder: {name}", icon="folder", url=url_for("projects.folder_detail", folder_key=fid))
    return redirect(url_for("projects.projects_list"))


@projects_bp.post("/folders/<folder_id>/color")
def recolor_folder(folder_id):
    """Quick swatch-only color change from the folder's ⋯ menu — no modal,
    keeps the existing name untouched. Works for built-in folders
    (Master/Weekly/Manual) as well as custom ones."""
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can recolor folders")
        return redirect(url_for("projects.projects_list"))
    color = request.form.get("color", "").strip()
    if folder_id in BUILTIN_FOLDER_IDS:
        if color:
            BUILTIN_FOLDERS[folder_id]["color"] = color
            flash("Folder color updated")
        return redirect(request.referrer or url_for("projects.projects_list"))
    folder = next((f for f in CUSTOM_FOLDERS if f["id"] == folder_id), None)
    if not folder:
        flash("Folder not found")
        return redirect(url_for("projects.projects_list"))
    if color:
        folder["color"] = color
        flash("Folder color updated")
    return redirect(request.referrer or url_for("projects.projects_list"))


@projects_bp.post("/folders/<folder_id>/duplicate")
def duplicate_folder(folder_id):
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can duplicate folders")
        return redirect(url_for("projects.projects_list"))
    folder = next((f for f in CUSTOM_FOLDERS if f["id"] == folder_id), None)
    if not folder:
        flash("Folder not found")
        return redirect(url_for("projects.projects_list"))
    fid = next_folder_id()
    CUSTOM_FOLDERS.append({"id": fid, "name": f"{folder['name']} (copy)", "color": folder.get("color", "#004BFF")})
    flash(f"Duplicated \"{folder['name']}\" — the copy is empty, ready to use")
    actor = prefs()["staff_name"]
    log_activity(actor, f"{actor} duplicated folder: {folder['name']}", icon="folder", url=url_for("projects.folder_detail", folder_key=fid))
    return redirect(url_for("projects.projects_list"))


@projects_bp.post("/folders/<folder_id>/empty")
def empty_folder(folder_id):
    """Move every project out of this folder (back to Manually added)
    without deleting the folder itself."""
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can move projects out of a folder")
        return redirect(url_for("projects.projects_list"))
    folder = next((f for f in CUSTOM_FOLDERS if f["id"] == folder_id), None)
    if not folder:
        flash("Folder not found")
        return redirect(url_for("projects.projects_list"))
    moved = 0
    for p in PROJECTS:
        if p.get("origin") == folder_id:
            p["origin"] = "manual"
            moved += 1
    if moved:
        flash(f"Moved {moved} project{'s' if moved != 1 else ''} out of \"{folder['name']}\" to Manually added")
        actor = prefs()["staff_name"]
        log_activity(actor, f"{actor} emptied folder: {folder['name']}", icon="folder", url=url_for("projects.folder_detail", folder_key=folder_id))
    else:
        flash("That folder is already empty")
    return redirect(url_for("projects.projects_list"))


@projects_bp.post("/folders/<folder_id>/edit")
def edit_folder(folder_id):
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can rename folders")
        return redirect(url_for("projects.projects_list"))
    name = request.form.get("name", "").strip()

    if folder_id in BUILTIN_FOLDER_IDS:
        b = BUILTIN_FOLDERS[folder_id]
        color = request.form.get("color", b["color"])
        if name:
            old_name = b["name"]
            b["name"] = name
            b["color"] = color
            flash("Folder updated")
            actor = prefs()["staff_name"]
            log_activity(actor, f"{actor} renamed folder \"{old_name}\" to \"{name}\"", icon="folder", url=url_for("projects.folder_detail", folder_key=folder_id))
        return redirect(url_for("projects.projects_list"))

    folder = next((f for f in CUSTOM_FOLDERS if f["id"] == folder_id), None)
    if not folder:
        flash("Folder not found")
        return redirect(url_for("projects.projects_list"))
    color = request.form.get("color", folder.get("color", "#004BFF"))
    if name:
        old_name = folder["name"]
        folder["name"] = name
        folder["color"] = color
        flash("Folder updated")
        actor = prefs()["staff_name"]
        log_activity(actor, f"{actor} renamed folder \"{old_name}\" to \"{name}\"", icon="folder", url=url_for("projects.folder_detail", folder_key=folder_id))
    return redirect(url_for("projects.projects_list"))


@projects_bp.post("/folders/<folder_id>/delete")
def delete_folder(folder_id):
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can delete folders")
        return redirect(url_for("projects.projects_list"))

    if folder_id in BUILTIN_FOLDER_IDS:
        # Built-in folders can't be removed as a category — "origin" always
        # has to fall back to something — so "Delete" here means wiping
        # everything inside instead: the projects in it and their tasks are
        # actually deleted, not moved out. The (now empty) category stays.
        b = BUILTIN_FOLDERS[folder_id]
        removed_ids = {p["id"] for p in PROJECTS if p.get("origin", "manual") == folder_id}
        removed = len(removed_ids)
        PROJECTS[:] = [p for p in PROJECTS if p["id"] not in removed_ids]
        TASKS[:] = [tk for tk in TASKS if tk["projectId"] not in removed_ids]
        MAIN_TASKS[:] = [m for m in MAIN_TASKS if m["projectId"] not in removed_ids]
        if removed:
            flash(f"Deleted {removed} project{'s' if removed != 1 else ''} and their tasks from \"{b['name']}\"")
        else:
            flash(f"\"{b['name']}\" was already empty")
        actor = prefs()["staff_name"]
        log_activity(actor, f"{actor} deleted all contents of folder: {b['name']}", icon="delete")
        return redirect(url_for("projects.projects_list"))

    folder = next((f for f in CUSTOM_FOLDERS if f["id"] == folder_id), None)
    if not folder:
        flash("Folder not found")
        return redirect(url_for("projects.projects_list"))
    # Projects that were grouped into this folder aren't deleted — they
    # just fall back to "Manually added" like any ungrouped project.
    moved = 0
    for p in PROJECTS:
        if p.get("origin") == folder_id:
            p["origin"] = "manual"
            moved += 1
    CUSTOM_FOLDERS[:] = [f for f in CUSTOM_FOLDERS if f["id"] != folder_id]
    msg = f"Folder \"{folder['name']}\" deleted"
    if moved:
        msg += f" — {moved} project{'s' if moved != 1 else ''} moved to Manually added"
    flash(msg)
    actor = prefs()["staff_name"]
    log_activity(actor, f"{actor} deleted folder: {folder['name']}", icon="delete")
    return redirect(url_for("projects.projects_list"))


@projects_bp.post("/<pid>/edit")
def edit_project(pid):
    project = project_of(pid)
    if not project:
        flash("Project not found")
        return redirect(url_for("projects.projects_list"))
    if not can_edit_project(project):
        flash("You can only edit projects you're assigned to")
        return redirect(url_for("projects.project_detail", pid=pid))

    name = request.form.get("name", "").strip()
    code = request.form.get("code", "").strip()
    if name:
        project["name"] = name
    if code:
        project["code"] = code
    project["category"] = request.form.get("category", "").strip() or project.get("category", "General")
    project["color"] = request.form.get("color", project.get("color", "#004BFF"))
    folder = request.form.get("folder")
    valid_folder_ids = {"master", "weekly", "manual"} | {f["id"] for f in CUSTOM_FOLDERS}
    if folder in valid_folder_ids:
        project["origin"] = folder

    if project.get("debtor"):
        _apply_case_details(project, request.form)
        _assign_case_task(project)

    flash("Project updated")
    actor = prefs()["staff_name"]
    log_activity(actor, f"{actor} updated case details for project: {project['name']}", icon="clock", project_id=pid)
    return redirect(url_for("projects.project_detail", pid=pid))


@projects_bp.post("/<pid>/pin")
def toggle_pin(pid):
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can pin projects")
        return redirect(request.referrer or url_for("dashboard.dashboard"))
    project = project_of(pid)
    if not project:
        flash("Project not found")
        return redirect(request.referrer or url_for("dashboard.dashboard"))
    project["pinned"] = not project.get("pinned")
    flash("Pinned to top" if project["pinned"] else "Unpinned")
    return redirect(request.referrer or url_for("dashboard.dashboard"))


@projects_bp.post("/<pid>/notify")
def notify_project(pid):
    redirect_to = request.form.get("redirect_to") or url_for("projects.project_detail", pid=pid)
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can send notifications")
        return redirect(redirect_to)
    project = project_of(pid)
    if not project:
        flash("Project not found")
        return redirect(redirect_to)

    recipient = project.get("pic") or project.get("collector")
    if not recipient:
        flash("This project has no PIC or collector assigned to notify")
        return redirect(redirect_to)

    sender = prefs()["staff_name"]
    notify(
        recipient,
        f"{sender} sent you a reminder about project \"{project['name']}\"",
        icon="alert",
    )
    log_activity(sender, f"{sender} notified {recipient} about project: {project['name']}", icon="bell", project_id=pid)
    flash(f"Notification sent to {recipient}")
    return redirect(redirect_to)


@projects_bp.post("/<pid>/delete")
def delete_project(pid):
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can delete projects")
        return redirect(url_for("projects.project_detail", pid=pid))
    project = project_of(pid)
    PROJECTS[:] = [p for p in PROJECTS if p["id"] != pid]
    TASKS[:] = [tk for tk in TASKS if tk["projectId"] != pid]
    flash("Project deleted")
    if project:
        actor = prefs()["staff_name"]
        log_activity(actor, f"{actor} deleted project: {project['name']}", icon="delete")
    return redirect(url_for("projects.projects_list"))


def _do_import_master(filename, stream):
    """Bulk-add projects from an uploaded .xlsx or .csv file.

    Two formats are recognized, auto-detected from the header row:

    - NPL "Master" sheet — a 'Debtor Name' column, one row per debtor/case.
      Every row becomes a project; Total Tagihan, PIC, DPD, due date, and
      the rest of the case's financials/status are carried over and shown
      on the project card and detail page.
    - Generic sheet — code, name, category, color columns (category and
      color are optional; color cycles through the app's palette).

    For .xlsx, the 'Master' sheet is read if present, otherwise the first
    (active) sheet.
    """
    lower_name = filename.lower()
    rows = []
    headers = set()
    wb = None
    try:
        if lower_name.endswith(".csv"):
            text = stream.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            headers = {(h or "").strip().lower() for h in (reader.fieldnames or [])}
            rows = [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]
        elif lower_name.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(stream, data_only=True)
            ws = wb["Master"] if "Master" in wb.sheetnames else wb.active
            it = ws.iter_rows()
            header = [str((c.value or "")).strip().lower() for c in next(it)]
            headers = set(header)
            for raw in it:
                if not any(c.value for c in raw):
                    continue
                row = {}
                for i in range(len(header)):
                    if i >= len(raw):
                        continue
                    cell = raw[i]
                    row[header[i]] = cell.value
                    # A cell can show a filename/label while actually linking
                    # out (e.g. a "Profiling"/"Doc Credit" cell whose text is
                    # "Profile.xlsx" but which opens a Google Drive file) —
                    # values_only=True would silently drop that link, so it's
                    # captured here under a separate key the field map reads.
                    if cell.hyperlink and cell.hyperlink.target:
                        row[header[i] + _LINK_KEY_SUFFIX] = cell.hyperlink.target
                rows.append(row)
        else:
            return "Unsupported file type — use .xlsx or .csv"
    except StopIteration:
        return "Could not read that file — the sheet looks empty (no header row found)"
    except Exception as ex:
        return f"Could not read that file — {ex}"

    try:
        added = _import_debtor_rows(rows) if "debtor name" in headers else _import_simple_rows(rows)
    except Exception as ex:
        return f"Could not process that file's rows — {ex}"

    msg = f"Imported {added} project(s)" if added else "No valid rows found — check the column headers"
    if wb is not None and "Interaksi" in wb.sheetnames:
        try:
            v_added, v_unmatched = _import_interaksi_tasks(wb["Interaksi"])
            if v_added:
                msg += f", {v_added} visit task(s) from Interaksi"
            if v_unmatched:
                msg += f" ({v_unmatched} visit debtor(s) had no matching project)"
        except Exception as ex:
            msg += f" (Interaksi sheet could not be processed — {ex})"
    return msg


def _do_import_weekly(filename, stream):
    """Import the weekly monitoring workbook (Internal Coll / Write Off /
    Auction / PKPU sheets). One task per debtor per workstream is created or
    updated — re-uploading next week's file just refreshes the same tasks
    with the latest note instead of piling up duplicates.
    """
    if not filename.lower().endswith(".xlsx"):
        return "The weekly update needs an .xlsx file"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(stream, data_only=True)
    except Exception as ex:
        return f"Could not read that file — {ex}"

    try:
        created, updated, new_projects = _import_weekly_tasks(wb)
    except Exception as ex:
        return f"Could not process that file's rows — {ex}"
    msg = f"Weekly update: {created} new task(s), {updated} refreshed"
    if new_projects:
        msg += f", {new_projects} new project(s) created for debtors not seen before"
    if created == 0 and updated == 0 and new_projects == 0:
        msg += " — no matching sheets found (expected one of: " + ", ".join(_WEEKLY_SHEETS.keys()) + ")"
    return msg



def _save_import_copy(kind, original_name, raw_bytes):
    """Persist a copy of the raw uploaded workbook to UPLOAD_DIR and record
    it in IMPORTS, so it can be reopened later from the Projects page
    instead of only seeing the data it produced."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    import_id = next_import_id()
    ext = os.path.splitext(original_name)[1].lower() or ".xlsx"
    stored_name = f"{import_id}{ext}"
    with open(os.path.join(UPLOAD_DIR, stored_name), "wb") as f:
        f.write(raw_bytes)
    record = {
        "id": import_id, "kind": kind, "filename": original_name,
        "stored_name": stored_name, "uploaded_by": prefs()["staff_name"],
        "uploaded_at": TODAY.isoformat(),
    }
    IMPORTS.append(record)
    return record


@projects_bp.post("/import")
def import_projects():
    """Single import entry point. The 'kind' field (from the Import Excel
    modal) picks whether the file is the Master data sheet (new/updated
    projects) or a Weekly update workbook (refreshes tasks)."""
    if not is_manager_or_admin():
        flash("Only Managers and Administrators can import projects")
        return redirect(url_for("projects.projects_list"))

    kind = request.form.get("kind", "master")

    if kind == "single":
        workstream = request.form.get("workstream", "")
        text = request.form.get("paste_text", "")
        single_file = request.files.get("single_file")
        touched_ids = []
        msg = _do_import_single_case(workstream, text, single_file, touched_out=touched_ids)
        flash(msg)
        if not msg.startswith(("Paste ", "Choose ", "Could not")):
            actor = prefs()["staff_name"]
            # A single-case update always touches exactly one debtor's
            # project — link the activity entry straight to it.
            touched_pid = touched_ids[0] if len(set(touched_ids)) == 1 else None
            log_activity(
                actor, f"{actor} logged a single-case update ({dict(_SINGLE_CASE_WORKSTREAMS).get(workstream, workstream)})",
                icon="import", project_id=touched_pid,
            )
        return redirect(url_for("projects.projects_list"))

    file = request.files.get("file")
    if not file or not file.filename:
        flash("Choose a file to import")
        return redirect(url_for("projects.projects_list"))

    original_name = file.filename
    display_name = re.sub(r"\.(xlsx|csv)$", "", original_name, flags=re.IGNORECASE)

    # Read the upload once into memory: one copy is saved to disk verbatim
    # so it can be downloaded later, another (fresh BytesIO, since openpyxl
    # consumes the stream) is handed to the parser.
    raw_bytes = file.read()
    msg = (
        _do_import_weekly(original_name, io.BytesIO(raw_bytes)) if kind == "weekly"
        else _do_import_master(original_name, io.BytesIO(raw_bytes))
    )
    flash(msg)
    _save_import_copy(kind, original_name, raw_bytes)

    actor = prefs()["staff_name"]
    kind_phrase = "weekly update excel" if kind == "weekly" else "excel"
    # A master/weekly import can create or touch many projects at once, so
    # there's no single project to link to — send the click to the folder
    # ("Weekly" or "Master data") that collects everything it just
    # imported instead.
    folder_key = "weekly" if kind == "weekly" else "master"
    log_activity(
        actor, f"{actor} imported {kind_phrase} called {display_name}", icon="import",
        url=url_for("projects.folder_detail", folder_key=folder_key),
    )
    return redirect(url_for("projects.projects_list"))


@projects_bp.get("/imports/<import_id>/download")
def download_import(import_id):
    """Send back the original spreadsheet exactly as it was uploaded."""
    record = next((imp for imp in IMPORTS if imp["id"] == import_id), None)
    if not record:
        flash("That file is no longer available")
        return redirect(url_for("projects.projects_list"))
    return send_from_directory(
        UPLOAD_DIR, record["stored_name"], as_attachment=True, download_name=record["filename"],
    )


def _fmt_xl_value(v):
    """openpyxl hands back whole numbers as floats (1 -> 1.0), which is
    correct data but reads oddly in a spreadsheet-style view — Excel itself
    displays a cell formatted as "1" not "1.0". Only collapse the .0 for
    values that are actually integral; a real decimal like 4.5 is untouched."""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _blank_if_whitespace(v):
    """Cells that only contain whitespace (a stray space left over from
    someone "clearing" a cell in Excel by hitting spacebar) are not real
    data. Left as-is, a lone " " counts as a truthy value everywhere it's
    compared — the "—" placeholder never gets selected in dropdown cells,
    and it won't match any option, so the cell renders as a select with an
    invisible space for a label instead of the intended blank state.
    Normalize those to a true empty string so blank means blank."""
    if isinstance(v, str) and v.strip() == "":
        return ""
    return v


def _read_import_sheets(record):
    """Load every sheet of an uploaded workbook (or the single table of a
    .csv) into plain header/rows data, for showing the file's contents
    inline instead of making people download it to look.

    Each row carries its original 0-based row index from the *unfiltered*
    file (i.e. the position openpyxl/csv would use to write it back), so
    edits made in the browser can be saved to the exact cell they came
    from even though blank rows are hidden from the view.
    """
    path = os.path.join(UPLOAD_DIR, record["stored_name"])
    sheets = []
    if record["stored_name"].lower().endswith(".csv"):
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            all_rows = list(csv.reader(f))
        kept = [(i, row) for i, row in enumerate(all_rows) if any(c.strip() for c in row)]
        if kept:
            header_idx, header_row = kept[0]
            # CSV has no concept of a hyperlink separate from the cell text,
            # so every cell's url is None here — kept as {value, url} dicts
            # anyway so the template doesn't need to branch by file type.
            # No merged cells either — that's an .xlsx-only concept.
            body = [{"idx": i, "cells": [{"value": _blank_if_whitespace(v), "url": None} for v in row]} for i, row in kept[1:]]
            sheets.append({
                "name": record["filename"], "header": header_row, "header_idx": header_idx, "rows": body,
                "merged": set(),
            })
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            raw_rows = [list(r) for r in ws.iter_rows()]
            kept = [(i, r) for i, r in enumerate(raw_rows) if any(c.value is not None and str(c.value).strip() != "" for c in r)]
            if not kept:
                continue
            header_idx, header_cells = kept[0]
            header = ["" if c.value is None else str(_fmt_xl_value(c.value)) for c in header_cells]
            # Every (row, col) — 0-based, matching the indices used below —
            # that falls inside a merged range in the original sheet, so the
            # template can flag it. Only the top-left cell of a merge holds
            # the value; the rest read as blank, which otherwise looks like
            # missing data rather than "this was merged in Excel".
            merged = set()
            for rng in ws.merged_cells.ranges:
                for r in range(rng.min_row - 1, rng.max_row):
                    for c in range(rng.min_col - 1, rng.max_col):
                        merged.add((r, c))
            body = [
                {
                    "idx": i,
                    "cells": [
                        {"value": "" if c.value is None else _blank_if_whitespace(_fmt_xl_value(c.value)),
                         "url": c.hyperlink.target if c.hyperlink and c.hyperlink.target else None}
                        for c in r
                    ],
                }
                for i, r in kept[1:]
            ]
            sheets.append({"name": name, "header": header, "header_idx": header_idx, "rows": body, "merged": merged})
    return sheets




@projects_bp.get("/imports/<import_id>/view")
def view_import(import_id):
    """Show everything in an uploaded workbook right in the app — every
    sheet, all rows, all at once — instead of sending the file back down
    to be opened in Excel."""
    record = next((imp for imp in IMPORTS if imp["id"] == import_id), None)
    if not record:
        flash("That file is no longer available")
        return redirect(url_for("projects.projects_list"))
    try:
        sheets = _read_import_sheets(record)
    except Exception:
        flash("Could not read that file")
        return redirect(url_for("projects.projects_list"))
    return render_template("import_view.html", view="projects", record=record, sheets=sheets, fixed_screen=True)


@projects_bp.post("/imports/<import_id>/save")
def save_import_edits(import_id):
    """Write cell edits made in the in-app spreadsheet view back to the
    stored file on disk, so the changes stick around next time it's opened."""
    record = next((imp for imp in IMPORTS if imp["id"] == import_id), None)
    if not record:
        return jsonify({"ok": False, "error": "That file is no longer available"}), 404

    payload = request.get_json(silent=True) or {}
    edits = payload.get("edits") or []
    if not isinstance(edits, list) or not edits:
        return jsonify({"ok": False, "error": "Nothing to save"}), 400

    path = os.path.join(UPLOAD_DIR, record["stored_name"])
    try:
        if record["stored_name"].lower().endswith(".csv"):
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))
            for e in edits:
                r, c, v = e.get("row"), e.get("col"), e.get("value", "")
                if r is None or c is None:
                    continue
                r, c = int(r), int(c)
                while len(rows) <= r:
                    rows.append([])
                while len(rows[r]) <= c:
                    rows[r].append("")
                rows[r][c] = v
            with open(path, "w", encoding="utf-8", newline="") as f:
                csv.writer(f).writerows(rows)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            for e in edits:
                sheet_name, r, c, v = e.get("sheet"), e.get("row"), e.get("col"), e.get("value", "")
                if sheet_name not in wb.sheetnames or r is None or c is None:
                    continue
                ws = wb[sheet_name]
                ws.cell(row=int(r) + 1, column=int(c) + 1, value=v)
            wb.save(path)
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)}), 500

    actor = prefs()["staff_name"]
    log_activity(actor, f"{actor} edited {record['filename']}", icon="import", url=url_for("projects.view_import", import_id=import_id))
    return jsonify({"ok": True})
